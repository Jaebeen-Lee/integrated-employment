# -*- coding: utf-8 -*-
import streamlit as st
import json
import io
import os
import pandas as pd
from datetime import datetime

# Excel & image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# 로컬 모듈 임포트 (동일 폴더에 employment_tax_credit_calc.py가 있어야 합니다)
from employment_tax_credit_calc import (
    CompanySize, Region, HeadcountInputs,
    load_params_from_json, calc_gross_credit,
    apply_caps_and_min_tax, calc_clawback, PolicyParameters
)

st.set_page_config(page_title="통합고용세액공제 계산기 (Pro, 메모리 로고)", layout="wide")

st.title("통합고용세액공제 계산기 · Pro (조특법 §29조의8)")
st.caption("로고를 메모리에서 직접 삽입합니다(임시파일X). 옵션에 따라 업로드한 PNG 로고를 세션에 저장해 계속 사용할 수 있습니다.")

# 세션 상태 초기화
if "saved_logo_png" not in st.session_state:
    st.session_state.saved_logo_png = None
if "saved_company_name" not in st.session_state:
    st.session_state.saved_company_name = None

with st.sidebar:
    st.header("1) 정책 파라미터")
    uploaded = st.file_uploader("시행령 기준 파라미터 JSON 업로드", type=["json"], accept_multiple_files=False)
    default_info = st.toggle("예시 파라미터 사용 (업로드 없을 때)", value=True)

    st.header("2) 보고서 옵션")
    company_name = st.text_input("회사/기관명 (머리글용)", value=st.session_state.saved_company_name or "(기관명)")
    logo_file = st.file_uploader("회사 로고 (PNG 권장)", type=["png"], accept_multiple_files=False)
    remember_logo = st.checkbox("이 로고를 계속 사용(세션에 저장)", value=True, help="브라우저 새로고침/재실행 시에도 유지됩니다. 앱/서버 재시작 시에는 초기화될 수 있습니다.")

    # 로고 세션 저장/복원
    logo_bytes = None
    if logo_file is not None:
        logo_bytes = logo_file.getvalue()
        if remember_logo:
            st.session_state.saved_logo_png = logo_bytes
    elif st.session_state.saved_logo_png is not None:
        logo_bytes = st.session_state.saved_logo_png

    # 회사명도 선택적으로 저장
    if company_name and remember_logo:
        st.session_state.saved_company_name = company_name

    params: PolicyParameters = None
    if uploaded is not None:
        try:
            cfg = json.load(uploaded)
            # 임시 파일 저장 후 로드 (모듈의 JSON 로더 재사용)
            tmp_path = "._tmp_params.json"
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False)
            params = load_params_from_json(tmp_path)
            os.remove(tmp_path)
            st.success("업로드한 파라미터를 불러왔습니다.")
        except Exception as e:
            st.error(f"파라미터 로딩 실패: {e}")
    elif default_info:
        # 데모용 기본 파라미터
        demo_cfg = {
            "per_head_basic": {
                "중소기업": {"수도권": 1200000, "지방": 1300000},
                "중견기업": {"수도권": 900000, "지방": 1000000},
                "대기업":   {"수도권": 600000, "지방": 700000}
            },
            "per_head_youth": {
                "중소기업": {"수도권": 1500000, "지방": 1600000},
                "중견기업": {"수도권": 1100000, "지방": 1200000},
                "대기업":   {"수도권": 800000,  "지방": 900000}
            },
            "per_head_conversion": 800000,
            "per_head_return_from_parental": 800000,
            "retention_years": {"중소기업": 3, "중견기업": 3, "대기업": 2},
            "max_credit_total": None,
            "min_tax_limit_rate": 0.07,
            "excluded_industries": ["유흥주점업", "기타소비성서비스업"]
        }
        tmp_path = "._tmp_params_demo.json"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(demo_cfg, f, ensure_ascii=False)
        params = load_params_from_json(tmp_path)
        os.remove(tmp_path)
        st.info("예시 파라미터를 사용 중입니다. (업로드 시 자동 대체)")

    st.divider()
    st.header("3) 기업 정보")
    size_label = st.selectbox("기업규모", [s.value for s in CompanySize], index=0, help="중소/중견/대기업 선택")
    region_label = st.selectbox("지역", [r.value for r in Region], index=1, help="사업장 소재지 기준")
    size = CompanySize(size_label)
    region = Region(region_label)

    st.divider()
    st.header("4) 사후관리 옵션")
    clawback_options = {
        "비례 추징 (감소율만큼)": "proportional",
        "전액 추징 (감소 발생 시 전체)": "all_or_nothing",
        "구간 추징 (감소율 구간별 단계)": "tiered"
    }
    selected_label = st.selectbox(
        "추징 방식 선택",
        list(clawback_options.keys()),
        index=0,
        help="감소율에 따라 추징액을 계산하는 방식을 선택합니다."
    )
    clawback_method = clawback_options[selected_label]

st.header("고용 인원 입력")
col1, col2, col3 = st.columns(3)

with col1:
    prev_total = st.number_input("전년 상시근로자 수", min_value=0, value=50, step=1)
    prev_youth = st.number_input("전년 청년등 상시근로자 수", min_value=0, value=10, step=1)
with col2:
    curr_total = st.number_input("당해 상시근로자 수", min_value=0, value=60, step=1)
    curr_youth = st.number_input("당해 청년등 상시근로자 수", min_value=0, value=14, step=1)
with col3:
    converted_regular = st.number_input("정규직 전환 인원 (해당연도)", min_value=0, value=2, step=1)
    returned_parental = st.number_input("육아휴직 복귀 인원 (해당연도)", min_value=0, value=1, step=1)

st.header("세액 한도/최저한세 옵션")
tax_before_credit = st.number_input(
    "세전세액(최저한세 적용 시 필요)",
    min_value=0, value=120_000_000, step=1,
    help="입력하지 않으면 최저한세 한도는 적용하지 않습니다."
)

st.divider()
run = st.button("계산하기", type="primary", disabled=(params is None))

if run:
    if params is None:
        st.error("파라미터(JSON)를 먼저 불러오세요.")
    else:
        heads = HeadcountInputs(
            prev_total=int(prev_total),
            curr_total=int(curr_total),
            prev_youth=int(prev_youth),
            curr_youth=int(curr_youth),
            converted_regular=int(converted_regular),
            returned_from_parental_leave=int(returned_parental),
        )
        gross = calc_gross_credit(size, region, heads, params)
        applied = apply_caps_and_min_tax(
            gross, params,
            tax_before_credit=int(tax_before_credit) if tax_before_credit else None
        )
        retention_years = params.retention_years[size]

        st.subheader("① 공제액 계산 결과")
        st.metric("총공제액 (최저한세/한도 적용 전)", f"{gross:,} 원")
        st.metric("적용 공제액 (최저한세/한도 적용 후)", f"{applied:,} 원")
        st.write(f"유지기간(사후관리 대상): **{retention_years}년**")

        # 다년 추징표
        st.subheader("② 사후관리(추징) 시뮬레이션 - 다년표")
        init_rows = []
        for yr in range(1, int(retention_years) + 1):
            init_rows.append({"연차": yr, "사후연도 인원": max(0, int(curr_total) - yr)})
        edited = st.data_editor(pd.DataFrame(init_rows), num_rows="dynamic")
        schedule = []
        for _, row in edited.iterrows():
            yidx = int(row["연차"])
            fol = int(row["사후연도 인원"])
            claw = calc_clawback(
                credit_applied=int(applied),
                base_headcount_at_credit=int(curr_total),
                headcount_in_followup_year=fol,
                retention_years_for_company=int(retention_years),
                year_index_from_credit=yidx,
                method=clawback_method,
            )
            schedule.append({
                "연차": yidx,
                "사후연도 인원": fol,
                "추징세액": int(claw)
            })
        schedule_df = pd.DataFrame(schedule).sort_values("연차").reset_index(drop=True)
        st.dataframe(schedule_df, use_container_width=True)
        total_clawback = int(schedule_df["추징세액"].sum())
        st.metric("추징세액 합계", f"{total_clawback:,} 원")

        # 엑셀 생성
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # 스타일
        title_font = Font(name="맑은 고딕", size=14, bold=True)
        header_fill = PatternFill("solid", fgColor="F2F2F2")
        thin = Side(style="thin", color="CCCCCC")
        border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        currency_style = NamedStyle(name="KRW")
        currency_style.number_format = '#,##0"원"'
        currency_style.alignment = right
        if "KRW" not in [ns.name for ns in wb.named_styles]:
            wb.add_named_style(currency_style)

        # 로고 삽입(메모리에서 직접)
        row_cursor = 1
        if logo_bytes is not None:
            try:
                pil_img = PILImage.open(io.BytesIO(logo_bytes))
                img = XLImage(pil_img)
                img.width = 140
                img.height = 40
                ws.add_image(img, "A1")
                row_cursor = 4
            except Exception as e:
                st.warning(f"로고 삽입 중 오류: {e}")

        # 타이틀/요약
        title_cell = ws.cell(row=row_cursor, column=1, value="통합고용세액공제 계산 결과")
        title_cell.font = title_font
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=6)
        ws.cell(row=row_cursor, column=7, value=f"작성일자: {datetime.now().strftime('%Y-%m-%d')}").alignment = right
        ws.cell(row=row_cursor+1, column=1, value=f"기관명: {company_name}")
        ws.cell(row=row_cursor+1, column=4, value=f"기업규모/지역: {size.value}/{region.value}")

        start = row_cursor + 3
        data = [
            ["항목", "값"],
            ["총공제액 (최저한세/한도 전)", int(gross)],
            ["적용 공제액 (최저한세/한도 후)", int(applied)],
            ["유지기간(년)", int(retention_years)],
            ["추징방식", clawback_method],
            ["추징세액 합계", total_clawback],
        ]
        for r_idx, row in enumerate(data, start=start):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        ws.cell(row=start+1, column=2).style = "KRW"
        ws.cell(row=start+2, column=2).style = "KRW"
        ws.cell(row=start+4, column=2).style = "KRW"

        for r in ws.iter_rows(min_row=start, max_row=start+len(data)-1, min_col=1, max_col=2):
            for cell in r:
                cell.border = border_all
                if cell.row == start:
                    cell.fill = header_fill
                    cell.alignment = center
                elif cell.column == 1:
                    cell.alignment = center
                else:
                    if cell.style != "KRW":
                        cell.alignment = right

        # 다년 추징표 시트
        ws2 = wb.create_sheet("Clawback Schedule")
        headers = ["연차", "사후연도 인원", "추징세액"]
        ws2.append(headers)
        for row in schedule:
            ws2.append([row["연차"], row["사후연도 인원"], row["추징세액"]])

        for cell in ws2[1]:
            cell.fill = header_fill
            cell.border = border_all
            cell.alignment = center
            cell.font = Font(bold=True)

        for r in range(2, 2 + len(schedule)):
            ws2.cell(row=r, column=1).alignment = center
            ws2.cell(row=r, column=2).alignment = right
            ws2.cell(row=r, column=3).style = "KRW"
            for c in range(1, 4):
                ws2.cell(row=r, column=c).border = border_all

        # 열 너비
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 26
        for col, w in zip(["A","B","C"], [10, 18, 18]):
            ws2.column_dimensions[col].width = w

        # 머리글(인쇄용)
        try:
            ws.header_footer.left_header = f"&L{company_name}"
            ws.header_footer.right_header = "&R통합고용세액공제 계산 결과"
            ws2.header_footer.left_header = f"&L{company_name}"
            ws2.header_footer.right_header = "&RClawback Schedule"
        except Exception:
            pass

        # 파라미터 시트
        ws3 = wb.create_sheet("Parameters")
        ws3.cell(row=1, column=1, value="Parameters (JSON)")
        ws3.cell(row=2, column=1, value=json.dumps({
            "per_head_basic": {k.value: {kk.value: v for kk, v in d.items()} for k, d in params.per_head_basic.items()},
            "per_head_youth": {k.value: {kk.value: v for kk, v in d.items()} for k, d in params.per_head_youth.items()},
            "per_head_conversion": params.per_head_conversion,
            "per_head_return_from_parental": params.per_head_return_from_parental,
            "retention_years": {k.value: v for k, v in params.retention_years.items()},
            "max_credit_total": params.max_credit_total,
            "min_tax_limit_rate": params.min_tax_limit_rate,
            "excluded_industries": params.excluded_industries,
        }, ensure_ascii=False, indent=2))

        wb.save(buffer)
        excel_name = f"tax_credit_result_pro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            label="엑셀 다운로드 (.xlsx, Pro 포맷)",
            file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            data=buffer.getvalue()
        )

else:
    st.info("좌측에서 파라미터(JSON)를 불러오고, 인원을 입력한 뒤 **계산하기**를 눌러주세요.")
